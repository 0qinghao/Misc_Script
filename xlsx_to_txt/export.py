import pandas as pd
from openpyxl import load_workbook

# 读取 Excel 文件
excel_file = 'Weekly Report - algorithm_rinlin20221226.xlsx'
# excel_file = 'Weekly Report - algorithm_rinlin20231225.xlsx'
# excel_file = 'Weekly Report - algorithm_rinlin20240624.xlsx'

# 使用 openpyxl 读取 sheet 名称的原始顺序
wb = load_workbook(excel_file, read_only=True)
sheet_names = wb.sheetnames

# 按 sheet 原始顺序倒序排序
sheet_names.reverse()

# 使用 pandas 读取所有 sheet
sheets = pd.read_excel(excel_file, sheet_name=None)

# 创建一个新的文本文件
with open('周报内容22.txt', 'w', encoding='utf-8') as file:
    # with open('周报内容23.txt', 'w', encoding='utf-8') as file:
    # with open('周报内容24.txt', 'w', encoding='utf-8') as file:
    # 遍历倒序排列的 sheet
    for sheet_name in sheet_names:
        data = sheets[sheet_name]
        # 写入 sheet 名称
        file.write(f'Sheet: {sheet_name}\n')

        # 遍历每一行，查找序号为1到5的格子，并提取右边一个格子的内容
        for index, row in data.iterrows():
            for col in range(len(row) - 1):
                if row[col] in [1, 2, 3, 4, 5]:
                    # 提取右边一个格子的内容
                    content = row[col + 1]
                    # 检查右侧格子是否有内容
                    if pd.notna(content):
                        file.write(f'{row[col]}: {content}\n')

        file.write('\n')

print("工作内容已提取并写入 '周报内容.txt' 文件中。")
